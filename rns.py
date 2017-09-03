import time
import sys
import os
import re
import getpass
import fnmatch
import win32api
import datetime
import sqlite3 as lite
import pandas as pd
import pyodbc
import xlwt
import glob

from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding('utf-8')
ned_req_file = 'MAS206545138B - Alexandra Village Pty Ltd.xlsx'
study_template = 'rns_template.xlsm'
template = os.path.abspath(study_template)
study_info = {}


def get_user():
    """ Get Planner's name from Windows domain. """
    username = getpass.getuser()
    fullname = win32api.GetUserNameEx(3)
    print(username, fullname)


def get_info(ned_req_file):
    """ Get info from the request file. """
    wb = load_workbook(filename=ned_req_file)
    sheet_names = wb.get_sheet_names()
    cn_sheet = wb.get_sheet_by_name('1. Customer & Tech Info')
    boq_sheet = wb.get_sheet_by_name('4. BOQ')
    
    for i in range(11, 50):
        label = cn_sheet.cell(row=i, column=1).value or ''
        info = cn_sheet.cell(row=i, column=2).value or ''
        if ':' in label:
            study_info[label[:-1]] = info
    nmd = re.sub('[^0-9]','', study_info['Existing Notified Demand'])
    new_nmd = re.sub('[^0-9]','', study_info['New NMD'])
    new_load = int(new_nmd) - int(nmd)
    study_info['New Load'] = new_load
    trfr = study_info['Mini-sub/Pole No.']
    for row_index in range(1, boq_sheet.max_row):
        transformer = ''
        try:
            if 'TFR' in boq_sheet.cell(row=row_index, column=2).value:
                transformer = boq_sheet.cell(row=row_index, column=2).value
                if 'kVA' in transformer:
                    transformer = str(transformer).split()[1]
                    transformer = int(re.sub('[^0-9]','', transformer))
                    study_info['Transformer'] = transformer        
            if 'CT+VT' in boq_sheet.cell(row=row_index, column=2).value:
                #transformer = boq_sheet.cell(row=row_index, column=2).value
                if re.sub('[^0-9]','', trfr[2]) <= 500:
                    transformer = '500 BULK'
                elif study_info['New Load'] >= 1000:
                    transformer = '1000 BULK'
                else:
                    transformer = '<1000 BULK'                    
                
                study_info['Transformer'] = transformer
        except TypeError:
            pass

def feeder_info():
    """ Get feeder info from heatmap. """
    db_filename = 'heatmap.db'
    con = lite.connect(db_filename)
    feeder = study_info['Feeder'].upper()[:-1]
    substation = study_info['Substation'].upper()
    fdrs = con.execute('SELECT "Feeder" from test').fetchall()
    for fdr in fdrs:
        try:
            if feeder in fdr[0].decode('utf-8'):
                feeder = fdr[0].decode('utf-8')
        except:
            pass
    feeder_info = con.execute('SELECT "Feeder","Substation","Voltage Class",\
                  "Peak Demand (kVA)","Installed Capacity (kVA)",\
                  "Backbone Conductors","Voltage (kV)","Senior Engineer",\
                  "Respective Manager" from test WHERE Feeder = ?', (feeder,)
                              ).fetchall()
    feeder_info = [x for x in feeder_info[0]]
    backbone_con = feeder_info[5].split(';')[0::2]
    feeder_info[5] = ', '.join(str(e).replace('#', '') for e in backbone_con)
    return feeder_info

# def perform_study(feeder):


def mdb_query(mdb_file, pole_number):
    """ Get pole or transformer info from case file."""
    # set up some constants
    MDB = mdb_file
    DRV = '{Microsoft Access Driver (*.mdb)}'; 

    # connect to db
    conn = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))
    cur = conn.cursor()
    type_tables = [{'lines': 'TypLne'}, {'transformers': 'TypTr2'}]
    element_tables = ['ElmLne', 'ElmTr2']
    
    line_types = {}
    trfr_types = {}
    for type_table in type_tables:
        type, table = [(k,v) for k,v in type_table.items()][0]
     
        SQL = 'SELECT * FROM [{}];'.format(table)
        if 'lines' in type:
            df_lines = pd.read_sql(SQL, conn)
            conductors = df_lines[['ID', 'loc_name']].apply(tuple,
                                                            axis=1).tolist()
            for conductor in conductors:
                k, v = conductor
                head, sep, tail = v.partition('3')
                line_types[k] = head
        if 'transformers' in type:
            df_transformers = pd.read_sql(SQL, conn)
            transformers = df_transformers[['ID',
                                            'loc_name']].apply(tuple,
                                                               axis=1).tolist()
            for i in transformers:
                k, v = i
                trfr_types[k] = v.split()

    # query pole parameters
    df_con =  pd.read_sql('SELECT * FROM [{}];'.format(element_tables[0]),
                                                       conn)
    df_type = pd.read_sql('SELECT * FROM [{}];'.format(element_tables[1]),
                                                       conn)        
    trfr_conductors = df_con[df_con['desc'].str.contains(pole_number)]   
    trfr_conductor_id = list(set(trfr_conductors.typ_id.tolist()))[0]
    trfr_type = df_type[df_type['loc_name'].str.contains(pole_number)]
    trfr_type_id = trfr_type.typ_id.tolist()[0]
    return line_types[trfr_conductor_id], trfr_types[trfr_type_id]
            
    
def create_vbs():
    project_num = study_info['Project Number']
    project_name = study_info['Project Name']
    study_project_id = '{0} - {1}'.format(project_num, project_name)
    supp_volt = study_info['Customer Supply Voltage']
    feeder = study_info['Feeder'].upper()[:-1]
    trfr = study_info['Mini-sub/Pole No.']
    
    if 'k' in supp_volt:
        supp_volt = int(''.join(x for x in supp_volt if x.isdigit()))
    else:
        supp_volt = int(''.join(x for x in supp_volt if x.isdigit()))/1000
    
    if supp_volt == '0.23':
        study_info['Connection Phase'] = 'Single Phase'
    else:
        study_info['Connection Phase'] = 'Three Phase'
    
    source_voltage = rmd_sort.source(rmd_file)
    power_factor = 0.98
    
    
    revision = '0.0'
    
    study_date = time.strftime("%Y%m%d")
    study_report = 'RNS_{0} - {1}_{2}_{3}.xlsx'.format(project_num,
                                                       project_name,
                                                       study_date,
                                                       revision)
    study_template = '''
    Set objExcel = CreateObject("Excel.Application") 
    objExcel.Visible = True 
    Set objWorkbook = objExcel.Workbooks.Open("$template")
    objExcel.Cells(9,2).Value = "$substation" 
    objExcel.Cells(5,2).Value = "$project_name"
    objExcel.Cells(7,2).Value = "$project_id"
    objExcel.Cells(11,2).Value = "$feeder_name"
    objExcel.Cells(13,2).Value = "$connection_point"
    objExcel.Cells(15,2).Value = "$feeder_class"
    objExcel.Cells(17,2).Value = "$connection_voltage"
    objExcel.Cells(19,2).Value = "$backbone_conductor"
    objExcel.Cells(22,2).Value = "$existing_trfr"
    objExcel.Cells(23,2).Value = "$new_trfr"
    objExcel.Cells(23,5).Value = "$new_load"
    objExcel.Cells(24,2).Value = "$connection_phase"
    objExcel.Cells(24,5).Value = "$connection_voltage"
    objExcel.Cells(27,2).Value = "$feeder_peak_load"
    objExcel.Cells(27,5).Value = "$feeder_voltage"
    objExcel.Cells(28,2).Value = "$powerfactor"
    objExcel.Cells(28,5).Value = "$source_volatge_pu" '%
    objExcel.Cells(30,2).Value = "$backbone_conductor"
    objExcel.Cells(31,2).Value = "$t_off_conductor"
    objExcel.Cells(36,5).Value = "$backfeeding_capacity"
    objExcel.Cells(39,2).Value = "$connection_voltage"
    objExcel.Cells(73,2).Value = "$user"
    objExcel.Cells(75,2).Value = "$senior"
    objExcel.Cells(77,2).Value = "$manager"
    objWorkbook.SaveAs "$study_report" 
    objWorkbook.Close 
    objExcel.Quit
    objExcel = Nothing
    Set objWorkbook = Nothing'''

get_info(ned_req_file)

print(study_info)
